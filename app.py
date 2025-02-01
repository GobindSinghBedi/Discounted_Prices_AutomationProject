# This project works on multiple excel workbooks to automate editing of these workbooks
# It corrects the price to discounted price by reducing it by 10%
# It also creates a bar chart

# For working on workbook -
import openpyxl as xl

# For working on chart -
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):  # 2,3,4,...,101
        cell = sheet.cell(row, 3)
        discounted_price = cell.value * 0.9  # Correcting the price by decreasing it by 10%
        discounted_price_cell = sheet.cell(row, 4)
        discounted_price_cell.value = discounted_price

    # Selecting the values for the BarChart
    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)
    labels = Reference(
        sheet, min_row=2, max_row=sheet.max_row, min_col=2, max_col=2)

    # Configuring and adding the BarChart to sheet
    chart = BarChart()
    chart.add_data(values)
    chart.set_categories(labels)
    chart.title = "Discounted Prices"
    chart.x_axis.title = "Product"
    chart.y_axis.title = "Price (in $)"
    sheet.add_chart(chart, "f2")

    wb.save(filename)


process_workbook("transactions.xlsx")
